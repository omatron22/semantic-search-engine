import sys
from openpyxl import load_workbook


def extract_text_from_xlsx(xlsx_path):
    """Extract text from Excel file by reading all sheets and rows."""
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        lines = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"Sheet: {sheet_name}")
            for row in ws.iter_rows(values_only=True):
                cells = [str(cell) if cell is not None else "" for cell in row]
                if any(cells):
                    lines.append(" | ".join(cells))

        wb.close()
        return "\n".join(lines).strip()
    except Exception as e:
        return f"Error reading XLSX: {str(e)}"


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python parse_xlsx.py <path_to_xlsx>")
        sys.exit(1)

    text = extract_text_from_xlsx(sys.argv[1])
    print(text)
